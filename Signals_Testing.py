import pandas as pd
import numpy as np
import datetime as dt
import time as time
import os

import arch.unitroot as UnitRoot
import matplotlib.pyplot as plt
from statsmodels.tsa.stattools import acf,pacf, adfuller
from operator import itemgetter
from sklearn import linear_model as LM
from sklearn.decomposition import PCA
from openpyxl import load_workbook
import openpyxl

#Takes a df of executeable prices and calcs the triple barrier label returns for each row
def getTripleBarrierLabel(data, lookback, time_limit, upper_std, lower_std):
    #@FORMAT: data = df(exec_prices,  index=dates)
    #time_limit is how long the algo waits before calculating the return. This will calc if neither upper bound or lower bound was reached.
    data_out = data.copy()
    for i in range(lookback, data_out.shape[0]):
        block = data_out.iloc[0, (i-lookback):i]
        st_dev = block.std()
        assert type(st_dev) == float, "std is not a number, check again dude"
        end = min(i + time_limit, data_out.shape[0]) #Walk forward is min of time limit or the end of the array
        for t in range(i, end):
            ret = (data_out.iloc[0, t] - data_out.iloc[0, i]) / data_out.iloc[0, i]
            if ret >= st_dev * upper_std or ret <= st_dev * lower_std:
                data_out.ix['triple_label', i] = ret
                break
        data_out.ix['triple_label', i] = (data_out.iloc[0, t] -  data_out.iloc[0, i]) / data_out.iloc[0, i]  #Neither upper or lower board was achieved, calc the time_limit return

    #@RETURN: data = df(prices, triple_label, index=dates)
    return data_out

def getNextExecutionLevels(data):
    #@FORMAT: data = df(prices,... index=dates)
    data_out = data.copy()
    for i in range(0, data.shape[0]):
        buy_found = False
        sell_found = False
        for ii in range(i, data.shape[0]-1):
            if not buy_found:
                if data.iloc[ii+1,0] < data.iloc[ii,0]:
                    data_out.ix[i, 'next_buy'] = data.iloc[ii,0]
                    buy_found = True

            if not sell_found:
                if data.iloc[ii+1,0] > data.iloc[ii,0]:
                    data_out.ix[i, 'next_sell'] = data.iloc[ii,0]
                    sell_found = True

            if buy_found and sell_found:
                break
    #@RETURN: data = df(prices, ..., next_buy, next_sell, index=dates)
    return data_out

#Applies 'fcn' to every block. It doesn't roll for every datapoint
def rolling_block_data_fcn(data,fcn,gap=5,*args,**kwargs):
    #@FORMAT: data = df(data,index=dates)
    dates = data.index.values
    values = data.values
    out = []
    out.append([dates[0],0])
    for i in range(0,values.shape[0],gap):
        block_values = values[i:i+gap]
        stat = fcn(block_values,**kwargs)
        out.append([dates[i],stat])
    return out

#Applies 'fcn' to every datapoint for a given lookback
def rolling_data_fcn(data,fcn,gap=5,*args,**kwargs):
    #@FORMAT: data = df(data,index=dates)
    dates = data.index.values
    values = data.values
    out = []
    out.append([dates[0],0])
    for i in range(0,values.shape[0],1):
        block_values = values[i:i+gap]
        stat = fcn(block_values,**kwargs)
        out.append([dates[i],stat])
    return out

def rolling_data_fcn2(data,fcn,gap=5,*args,**kwargs):
    #@FORMAT: data = np.array
    out = np.empty([data.shape[0],1])
    out[:] = np.nan
    for i in range(0,data.shape[0],1):
        block_values = data[i:i+gap]
        stat = fcn(block_values,**kwargs)
        out.append([stat])
    return out



#Applies each 'fcn' to a rolling block of data
def getRollingTraits(data,fcn_list,gap=5,*args,**kwargs):
    #@FORMAT: data = df(price,index=dates) - Only Price Column
    #@FYI: Each 'fcn' needs to be able to take 1D numpy array and outputs one value

    values = data.values
    traits_data = pd.DataFrame(index=data.index.values, columns=range(0, len(fcn_list)))

    for i in range(gap,traits_data.shape[0],1):
        block_values = values[(i-gap):i,:]
        stats = []
        for fcn in fcn_list:
            stats.append(fcn(block_values))
        traits_data.loc[traits_data.index.values[i]] = stats

    #@RETURNS: traits_df = df(traits,index=dates)
    return traits_data


#WRITING TO EXCEL FUNCTIONS

#Makes a new excel sheet and writes to it
def write_new(datadf, path, tab="Sheet1"):
    WRITE_PATH = path
    writer = pd.ExcelWriter(WRITE_PATH, engine='xlsxwriter')
    datadf.to_excel(writer, sheet_name=tab)
    writer.save()

#Takes existing excel file and writes to specified sheet
def write(datadf, path, tab="Sheet1"):
    WRITE_PATH = path
    book = load_workbook(path)
    writer = pd.ExcelWriter(WRITE_PATH, engine='openpyxl')
    writer.book = book
    datadf.to_excel(writer, sheet_name=tab)
    writer.save()

#Takes existing excel file and overwrites the existing sheet
def write_overwritesheet(datadf, path, tab="Sheet1"):
    #Workbook must have more than 1 worksheets
    WRITE_PATH = path

    #First delete the sheet
    workbook = openpyxl.load_workbook(WRITE_PATH)
    ws = workbook[tab]
    workbook.remove(ws)
    workbook.save(WRITE_PATH)

    book = load_workbook(path)
    writer = pd.ExcelWriter(WRITE_PATH, engine='openpyxl')
    writer.book = book
    datadf.to_excel(writer, sheet_name=tab)
    writer.save()

